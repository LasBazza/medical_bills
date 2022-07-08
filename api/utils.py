from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import ColumnNames
from .exceptions import UnknownClientColumns


class ClientOrganizationExcelParser:

    def parse_excel(self, file: str, sheet_title: str) -> list[dict]:
        """Return list of instance dictionaries from specified excel sheet"""

        wb = load_workbook(filename=file)
        sheet = wb[sheet_title]

        field_names = self._get_fields_names_from_table(sheet)

        items = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            item = dict()
            for idx in range(len(row)):
                if row[idx] is not None:
                    item[field_names[idx]] = row[idx]
            if item:
                items.append(item)

        return items

    @staticmethod
    def _get_fields_names_from_table(sheet: Worksheet) -> list[str]:
        """Read first row of table and return list of fileds names for next usage as instance dictionary keys"""

        field_names = []
        for row in sheet.iter_rows(max_row=1, values_only=True):
            for cell in row:
                if cell is not None:
                    field_names.append(cell)

        return field_names


class BillsParser:

    def parse_excel(self, file: str, sheet_title: str):

        wb = load_workbook(filename=file)
        sheet = wb[sheet_title]

        client_columns = self._recognize_client(sheet)

        # following steps of parsing are not implemented

    @staticmethod
    def _recognize_client(sheet):
        client_names_columns = []
        for instance in ColumnNames.objects.all():
            client_names_columns.append(instance.client_name_column)

        for row in sheet.iter_rows(max_row=1, values_only=True):
            for cell in row:
                if cell is not None:
                    if cell in client_names_columns:
                        return ColumnNames.objects.get(client_name_column=cell)
                    else:
                        raise UnknownClientColumns

    @staticmethod
    def _normalize_bill_number(value: Any) -> int:
        return int(value)

    @staticmethod
    def _normalize_sum(value: Any) -> float:
        return float(value)

